"""Inspect results from check discovery rules

This script allows the user to inspect results from multiple `tpframework checkdiscoveryrules` executions and
build a combined excel.

This script requires the file
    * ./resources/inspect_disc_rules_results.json

This json file specifies which are the sources of the check discovery rules results.
"""

from pathlib import Path
from datetime import datetime
import csv
from openpyxl import Workbook
import json

dirname = Path(__file__).parent.resolve()

# fetching data from config file
data_file = dirname / "resources/inspect_disc_rules_results.json"
if not data_file.is_file():
    print(f"ERROR - data file `{data_file}` does not exist. Please create it from the template file `{data_file}.template`.")
    exit(1)
with open(data_file, "r") as f:
    data = json.load(f)
rel_data_root_path = data["rel_data_root_path"]
data_sources = data["sources"]
data_root = (dirname / rel_data_root_path).resolve()

# date format
date_format = '%d/%m/%Y'

# safety check on dates
for d in data_sources:
    if not datetime.strptime(d['date'], date_format):
        print(
            f"ERROR - `{d}` does not comply with the date format `{date_format}`. Please correct it in the data file `{data_file}`.")
        exit(1)


def get_exception_message(e):
    if hasattr(e, 'message'):
        return e.message
    elif hasattr(e, 'msg'):
        return e.msg
    else:
        return str(e)


def extend_row_data(row: dict, date=None):
    if "instance_path" in row.keys() and not "instance_name" in row.keys():
        row["instance_name"] = row["instance_path"].split("/")[-1]
    row["date"] = date
    return row


def load_source(src: Path):
    src_data = []
    count = 0
    for sfpath in src.iterdir():
        if sfpath.suffix == ".csv":
            count += 1
            with open(sfpath, newline='') as sf:
                csv_data = csv.DictReader(sf, delimiter=',')
                for row in csv_data:
                    src_data.append(row)
    if count == 0:
        print(f"WARNING - No source file found in {src}")
    elif count == 1:
        print(f"One source file found and loaded from {src}")
    elif count > 1:
        print(f"WARNING - More than one source files found in {src}")
    return src_data


def inspect():
    print(f"Loading sources...")
    for entry_data in data_sources:
        try:
            src_data = load_source(data_root / entry_data["rel_path"])
            status = "success"
        except Exception as e:
            src_data = {}
            status = "error: " + get_exception_message(e)
        finally:
            entry_data["data"] = src_data
            entry_data["status"] = status

    print(f"Cluster source data by language and date...")
    cdata = {}
    for entry_data in data_sources:
        date = entry_data["date"]
        for tpi in entry_data["data"]:
            lang = tpi["language"]
            tpi = extend_row_data(tpi, date=date)
            if lang not in cdata.keys():
                cdata[lang] = {
                    date: [tpi]
                }
            elif date not in cdata[lang].keys():
                cdata[lang][date] = [tpi]
            else:
                cdata[lang][date].append(tpi)
    #
    print("Create summary sheet...")
    wb = Workbook()
    summary = wb.active
    summary.title = "disc_rules_summary"
    summary.cell(row=2, column=1, value="counting")
    summary.cell(row=3, column=1, value="successful")
    summary.cell(row=4, column=1, value="unsuccessful")
    summary.cell(row=5, column=1, value="error")
    summary.cell(row=6, column=1, value="missing")
    summary.cell(row=7, column=1, value="total")
    summary.cell(row=8, column=1, value="discrepancies")
    col = 2
    for lang in cdata:
        sorted_dates = [datetime.strftime(d, date_format) for d in
                        sorted([datetime.strptime(d, date_format) for d in cdata[lang]])]
        for date in sorted_dates:
            counting = len(cdata[lang][date])
            successful = len([d for d in cdata[lang][date] if d["successful"] == "yes"])
            unsuccessful = len([d for d in cdata[lang][date] if d["successful"] == "no"])
            error = len([d for d in cdata[lang][date] if d["successful"] == "error"])
            missing = len([d for d in cdata[lang][date] if d["successful"] == "missing"])
            total = successful + unsuccessful + error + missing
            summary.cell(row=1, column=col, value=f"{lang} - {date}")
            summary.cell(row=2, column=col, value=counting)
            summary.cell(row=3, column=col, value=successful)
            summary.cell(row=4, column=col, value=unsuccessful)
            summary.cell(row=5, column=col, value=error)
            summary.cell(row=6, column=col, value=missing)
            summary.cell(row=7, column=col, value=total)
            summary.cell(row=8, column=col, value=counting - total)
            col += 1
    #
    print(f"Create a sheet per language...")
    header_prefix = [
        "instance_name", # extension
        "date", # extension
        "pattern_id",
        "instance_id",
        "instance_path",
        "pattern_name",
        "language",
        "discovery_rule"
    ]
    header_suffix = [
        "diff",
        "deteriorate"
    ]
    for lang in cdata:
        columns = {}
        added = {}
        sheet = wb.create_sheet(title=f"{lang}-disc_rules")
        # add prefix headers
        for i, h in enumerate(header_prefix):
            columns[h] = i + 1
            sheet.cell(row=1, column=i+1, value=h)
        # add header dates
        sorted_dates = [datetime.strftime(d, date_format) for d  in sorted([datetime.strptime(d, date_format) for d in cdata[lang]])]
        col = 0
        for i, date in enumerate(sorted_dates):
            col = len(header_prefix)+1+i
            columns[date] = col
            sheet.cell(row=1, column=col, value=date)
        # add suffix headers
        for i, h in enumerate(header_suffix):
            columns[h] = col + i + 1
            sheet.cell(row=1, column=col+i+1, value=h)
        #
        irow = 2
        for date in cdata[lang]:
            col = columns[date]
            for tpi in cdata[lang][date]:
                if tpi["instance_name"] and tpi["instance_name"] in added.keys():
                    # add in existing row
                    erow = added[tpi["instance_name"]]
                    if sheet.cell(row=erow, column=col).value:
                        print(f"WARNING - {tpi['instance_name']} - a value for that experiment date exists already ({sheet.cell(row=1, column=col).value})...the new one ({tpi['successful']}) will be ignored")
                    else:
                        sheet.cell(row=erow, column=col, value=tpi["successful"])
                else:
                    # new row
                    for i in range(1, len(header_prefix)+1):
                        sheet.cell(row=irow, column=i, value=tpi[header_prefix[i - 1]])
                    sheet.cell(row=irow, column=col, value=tpi["successful"])
                    added[tpi["instance_name"]] = irow
                    irow += 1

        if len(sorted_dates) > 1:
            l_date_col = [columns[d] for d in sorted_dates]
            # compute suffix columns values
            for i in range(2, sheet.max_row):
                # diff
                diff = not all(sheet.cell(row=i, column=col1).value==sheet.cell(row=i, column=col2).value
                           for col1 in l_date_col for col2 in l_date_col if col1!=col2)
                sheet.cell(row=i, column=columns["diff"], value=diff)
                # deterioration
                last = sheet.cell(row=i, column=l_date_col[-1]).value
                l_prev = [sheet.cell(row=i, column=col).value for col in l_date_col[:-1]]
                deter = (last!="yes" and any(prev=="yes" for prev in l_prev))
                sheet.cell(row=i, column=columns["deteriorate"], value=deter)

    now: datetime = datetime.now()
    now_str = now.strftime("%Y%m%d_%H%M%S")
    ofname = f"{now_str}-disc_rules_status.xlsx"
    wb.save(data_root / ofname)
    print(f"Output excel file: {data_root / ofname}")


if __name__ == "__main__":
    try:
        inspect()
    except KeyboardInterrupt:
        print("Interrupting...")
